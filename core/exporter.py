import csv
from datetime import datetime


def _chosen_allowable(record, is_rural=False):
    """Return the display allowable amount based on rural flag.

    Priority:
    - rural=True  → allowable_r, fallback to allowable_nr if r is None/0
    - rural=False → allowable_nr, fallback to legacy allowable
    """
    if is_rural:
        r = record.get("allowable_r")
        if r:
            return r
        # Fall back to NR when rural amount not available
    nr = record.get("allowable_nr")
    if nr is not None:
        return nr
    return record.get("allowable")


def export_to_csv(records, filepath, is_rural=False):
    if not records:
        return
    fieldnames = [
        "hcpcs_code", "description", "state_abbr", "year",
        "allowable_nr", "allowable_r", "allowable", "modifier", "data_source",
    ]
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for r in records:
            row = {k: r.get(k, "") for k in fieldnames}
            chosen = _chosen_allowable(r, is_rural=is_rural)
            row["allowable"] = "#N/A" if chosen is None else chosen
            if row["allowable_nr"] is None:
                row["allowable_nr"] = "#N/A"
            if row["allowable_r"] is None:
                row["allowable_r"] = "#N/A"
            writer.writerow(row)


def export_to_excel(records, filepath, is_rural=False):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "HCPCS Fee Schedule"

    # Header
    headers = [
        "HCPCS Code", "Description", "State", "Year",
        "Allowable (NR)", "Allowable (R)", "Allowable ($)", "Modifier", "Source",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="003366")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    na_fill = PatternFill("solid", fgColor="FFF9C4")
    alt_fill = PatternFill("solid", fgColor="EEF2F7")
    for row_i, r in enumerate(records, 2):
        chosen = _chosen_allowable(r, is_rural=is_rural)
        is_na = chosen is None
        nr = r.get("allowable_nr")
        rv = r.get("allowable_r")
        values = [
            r.get("hcpcs_code", ""),
            r.get("description", ""),
            r.get("state_abbr", ""),
            r.get("year", ""),
            "#N/A" if nr is None else nr,
            "#N/A" if rv is None else rv,
            "#N/A" if is_na else chosen,
            r.get("modifier", "") or "",
            r.get("data_source", "") or "",
        ]
        for col_i, v in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            if is_na:
                cell.fill = na_fill
            elif row_i % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    widths = [12, 60, 8, 6, 14, 14, 14, 10, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    wb.save(filepath)


def export_to_pdf(records, filepath, is_rural=False):
    try:
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
    except ImportError:
        raise ImportError("reportlab is required for PDF export. Run: pip install reportlab")

    doc = SimpleDocTemplate(filepath, pagesize=landscape(letter),
                            topMargin=0.5 * inch, bottomMargin=0.5 * inch,
                            leftMargin=0.5 * inch, rightMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title = Paragraph("<b>VA HCPCS Fee Schedule Report</b>", styles["Title"])
    story.append(title)
    subtitle = Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}  |  {len(records):,} records",
        styles["Normal"],
    )
    story.append(subtitle)
    story.append(Spacer(1, 0.2 * inch))

    # Table data
    col_headers = ["HCPCS Code", "Description", "State", "Year", "Allowable ($)", "Modifier"]
    table_data = [col_headers]
    for r in records:
        chosen = _chosen_allowable(r, is_rural=is_rural)
        table_data.append([
            r.get("hcpcs_code", ""),
            (r.get("description", "") or "")[:80],
            r.get("state_abbr", ""),
            str(r.get("year", "")),
            "#N/A" if chosen is None else f"${chosen:,.2f}",
            r.get("modifier", "") or "",
        ])

    col_widths = [1.0 * inch, 4.5 * inch, 0.6 * inch, 0.6 * inch, 1.0 * inch, 0.8 * inch]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F7")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("WORDWRAP", (1, 1), (1, -1), True),
    ]))
    story.append(table)
    doc.build(story)
